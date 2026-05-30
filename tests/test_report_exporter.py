"""
ReportExporter 单元测试
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import io
import json
import zipfile
import pytest

# ---------------------------------------------------------------------------
# Dependency checks
# ---------------------------------------------------------------------------

try:
    from docx import Document as _DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl  # noqa: F401
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import reportlab  # noqa: F401
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Conditionally import functions (always import the stateless ones)
from report_exporter import (
    _normalize_diff_items,
    _risk_color_hex,
    _risk_color_rgb,
    _risk_label_cn,
    _category_label_cn,
    _direction_label_cn,
    export_json_api,
)


# ====================== 测试数据 ======================

@pytest.fixture
def empty_comparison_result():
    return {
        "numbers": {"matched": [], "missing_in_pdf": [], "extra_in_pdf": [], "has_diff": False},
        "dates": {"matched": [], "missing_in_pdf": [], "extra_in_pdf": [], "has_diff": False},
        "amounts_words": {"matched": [], "missing_in_pdf": [], "extra_in_pdf": [], "has_diff": False},
        "amounts_digits": {"matched": [], "missing_in_pdf": [], "extra_in_pdf": [], "has_diff": False},
        "percentages": {"matched": [], "missing_in_pdf": [], "extra_in_pdf": [], "has_diff": False},
    }


@pytest.fixture
def sample_comparison_result():
    return {
        "numbers": {
            "matched": [{"raw": "365", "normalized": 365.0, "context": "期限"}],
            "missing_in_pdf": [],
            "extra_in_pdf": [],
            "has_diff": False,
        },
        "dates": {
            "matched": [{"raw": "2024-01-15", "normalized": "2024-01-15"}],
            "missing_in_pdf": [],
            "extra_in_pdf": [],
            "has_diff": False,
        },
        "amounts_words": {
            "matched": [],
            "missing_in_pdf": [{"raw": "壹拾伍万元整", "context": "合同金额"}],
            "extra_in_pdf": [],
            "has_diff": True,
        },
        "amounts_digits": {
            "matched": [],
            "missing_in_pdf": [
                {"raw": "150000", "normalized": 150000.0, "keyword": "合同金额", "context": "合同金额为150000元", "phrase": "合同金额为150000元"},
            ],
            "extra_in_pdf": [],
            "has_diff": True,
        },
        "percentages": {
            "matched": [],
            "missing_in_pdf": [],
            "extra_in_pdf": [{"raw": "5%", "normalized": 0.05}],
            "has_diff": True,
        },
    }


@pytest.fixture
def sample_summary():
    return {
        "total_diffs": 3,
        "has_critical_diff": True,
        "diff_details": [
            {"type": "大写金额", "missing": 1, "extra": 0},
            {"type": "金额数字", "missing": 1, "extra": 0},
            {"type": "百分比", "missing": 0, "extra": 1},
        ],
    }


@pytest.fixture
def sample_word_text():
    return "合同编号：HT-2024-001\n合同金额：壹拾伍万元整\n签订日期：2024年1月15日"


@pytest.fixture
def sample_pdf_text():
    return "合同编号：HT-2024-001\n签订日期：2024年1月15日\n违约金比例：5%"


@pytest.fixture
def empty_summary():
    return {"total_diffs": 0, "has_critical_diff": False}


# ====================== 辅助函数测试 ======================

class TestNormalizeDiffItems:
    """差异项归一化测试"""

    def test_empty_result(self, empty_comparison_result):
        result = _normalize_diff_items(empty_comparison_result)
        assert "amounts_digits" in result
        assert "amounts_words" in result
        assert "dates" in result
        assert "numbers" in result
        assert "percentages" in result
        for v in result.values():
            assert v == []

    def test_missing_items_collected(self, sample_comparison_result):
        result = _normalize_diff_items(sample_comparison_result)
        # amounts_digits has 1 missing_in_pdf
        assert len(result["amounts_digits"]) == 1
        assert result["amounts_digits"][0]["direction"] == "missing_in_pdf"
        assert result["amounts_digits"][0]["raw"] == "150000"
        assert result["amounts_digits"][0]["risk"] == "high"

    def test_extra_items_collected(self, sample_comparison_result):
        result = _normalize_diff_items(sample_comparison_result)
        # percentages has 1 extra_in_pdf
        assert len(result["percentages"]) == 1
        assert result["percentages"][0]["direction"] == "extra_in_pdf"
        assert result["percentages"][0]["risk"] == "low"

    def test_normalized_items_have_required_fields(self, sample_comparison_result):
        result = _normalize_diff_items(sample_comparison_result)
        for cat_items in result.values():
            for item in cat_items:
                assert "raw" in item
                assert "direction" in item
                assert "risk" in item
                assert "category" in item


class TestRiskColorFunctions:
    """风险颜色函数测试"""

    def test_high_risk_color_hex(self):
        assert _risk_color_hex("high") == "D32F2F"

    def test_medium_risk_color_hex(self):
        assert _risk_color_hex("medium") == "F57C00"

    def test_low_risk_color_hex(self):
        assert _risk_color_hex("low") == "1976D2"

    def test_unknown_risk_color_hex(self):
        assert _risk_color_hex("unknown") == "666666"

    def test_high_risk_color_rgb(self):
        assert _risk_color_rgb("high") == (0xD3, 0x2F, 0x2F)

    def test_medium_risk_color_rgb(self):
        assert _risk_color_rgb("medium") == (0xF5, 0x7C, 0x00)

    def test_low_risk_color_rgb(self):
        assert _risk_color_rgb("low") == (0x19, 0x76, 0xD2)

    def test_unknown_risk_color_rgb(self):
        assert _risk_color_rgb("unknown") == (0x66, 0x66, 0x66)


class TestLabelFunctions:
    """标签函数测试"""

    def test_risk_label_high(self):
        assert _risk_label_cn("high") == "高风险"

    def test_risk_label_medium(self):
        assert _risk_label_cn("medium") == "中风险"

    def test_risk_label_low(self):
        assert _risk_label_cn("low") == "低风险"

    def test_risk_label_unknown(self):
        assert _risk_label_cn("unknown") == "unknown"

    @pytest.mark.parametrize("key,label", [
        ("amounts_digits", "金额数字"),
        ("amounts_words", "大写金额"),
        ("dates", "日期"),
        ("numbers", "数字"),
        ("percentages", "百分比"),
    ])
    def test_category_label(self, key, label):
        assert _category_label_cn(key) == label

    def test_category_label_unknown(self):
        assert _category_label_cn("unknown_key") == "unknown_key"

    def test_direction_label_missing(self):
        assert _direction_label_cn("missing_in_pdf") == "PDF 缺失"

    def test_direction_label_extra(self):
        assert _direction_label_cn("extra_in_pdf") == "PDF 多出"


# ====================== JSON API 导出测试 ======================

class TestExportJsonApi:
    """JSON API 导出测试"""

    def test_valid_json_output(self, sample_comparison_result, sample_summary,
                                sample_word_text, sample_pdf_text):
        result = export_json_api(sample_comparison_result, sample_summary,
                                  sample_word_text, sample_pdf_text)
        data = json.loads(result)
        assert isinstance(data, dict)

    def test_schema_version(self, sample_comparison_result, sample_summary,
                             sample_word_text, sample_pdf_text):
        result = export_json_api(sample_comparison_result, sample_summary,
                                  sample_word_text, sample_pdf_text)
        data = json.loads(result)
        assert data["schema_version"] == "2.0"

    def test_required_top_level_keys(self, sample_comparison_result, sample_summary,
                                      sample_word_text, sample_pdf_text):
        result = export_json_api(sample_comparison_result, sample_summary,
                                  sample_word_text, sample_pdf_text)
        data = json.loads(result)
        for key in ["schema_version", "timestamp", "metadata", "summary",
                     "differences", "full_text_diff", "statistics"]:
            assert key in data, f"Missing key: {key}"

    def test_metadata_contains_engine_info(self, sample_comparison_result, sample_summary,
                                            sample_word_text, sample_pdf_text):
        result = export_json_api(sample_comparison_result, sample_summary,
                                  sample_word_text, sample_pdf_text)
        data = json.loads(result)
        assert data["metadata"]["engine"] == "contract_comparator"
        assert data["metadata"]["engine_version"] == "2.0"

    def test_differences_are_list(self, sample_comparison_result, sample_summary,
                                   sample_word_text, sample_pdf_text):
        result = export_json_api(sample_comparison_result, sample_summary,
                                  sample_word_text, sample_pdf_text)
        data = json.loads(result)
        assert isinstance(data["differences"], list)

    def test_diff_records_have_required_fields(self, sample_comparison_result,
                                                sample_summary, sample_word_text,
                                                sample_pdf_text):
        result = export_json_api(sample_comparison_result, sample_summary,
                                  sample_word_text, sample_pdf_text)
        data = json.loads(result)
        for diff in data["differences"]:
            for field in ["id", "category", "category_label", "risk_level",
                          "direction", "direction_label", "value"]:
                assert field in diff, f"Missing field: {field}"

    def test_statistics_by_risk_level(self, sample_comparison_result, sample_summary,
                                       sample_word_text, sample_pdf_text):
        result = export_json_api(sample_comparison_result, sample_summary,
                                  sample_word_text, sample_pdf_text)
        data = json.loads(result)
        stats = data["statistics"]
        assert "by_risk_level" in stats
        assert "high" in stats["by_risk_level"]
        assert "medium" in stats["by_risk_level"]
        assert "low" in stats["by_risk_level"]

    def test_statistics_by_category(self, sample_comparison_result, sample_summary,
                                     sample_word_text, sample_pdf_text):
        result = export_json_api(sample_comparison_result, sample_summary,
                                  sample_word_text, sample_pdf_text)
        data = json.loads(result)
        assert "by_category" in data["statistics"]

    def test_statistics_by_direction(self, sample_comparison_result, sample_summary,
                                      sample_word_text, sample_pdf_text):
        result = export_json_api(sample_comparison_result, sample_summary,
                                  sample_word_text, sample_pdf_text)
        data = json.loads(result)
        by_dir = data["statistics"]["by_direction"]
        assert "missing_in_pdf" in by_dir
        assert "extra_in_pdf" in by_dir

    def test_empty_data_no_crash(self, empty_comparison_result, empty_summary):
        result = export_json_api(empty_comparison_result, empty_summary, "", "")
        data = json.loads(result)
        assert data["summary"]["total_differences"] == 0
        assert data["differences"] == []

    def test_total_differences_matches(self, sample_comparison_result, sample_summary,
                                        sample_word_text, sample_pdf_text):
        result = export_json_api(sample_comparison_result, sample_summary,
                                  sample_word_text, sample_pdf_text)
        data = json.loads(result)
        stats = data["statistics"]
        assert stats["total_differences"] == len(data["differences"])


# ====================== Redline DOCX 导出测试 ======================

@pytest.mark.skipif(not DOCX_AVAILABLE, reason="python-docx not installed")
class TestExportRedlineDocx:
    """Redline DOCX 导出测试"""

    def _get_export_func(self):
        from report_exporter import export_redline_docx
        return export_redline_docx

    def test_creates_valid_docx(self, sample_comparison_result, sample_summary,
                                 sample_word_text, sample_pdf_text):
        export_func = self._get_export_func()
        result = export_func(sample_comparison_result, sample_summary,
                              sample_word_text, sample_pdf_text)
        assert isinstance(result, bytes)
        assert len(result) > 0
        # DOCX is a ZIP file, should be loadable
        zf = zipfile.ZipFile(io.BytesIO(result))
        assert "[Content_Types].xml" in zf.namelist() or "word/document.xml" in zf.namelist()

    def test_empty_data_no_crash(self, empty_comparison_result, empty_summary):
        export_func = self._get_export_func()
        result = export_func(empty_comparison_result, empty_summary, "", "")
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_returns_bytes_object(self, sample_comparison_result, sample_summary,
                                   sample_word_text, sample_pdf_text):
        export_func = self._get_export_func()
        result = export_func(sample_comparison_result, sample_summary,
                              sample_word_text, sample_pdf_text)
        assert isinstance(result, bytes)


# ====================== Excel Diff 导出测试 ======================

@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
class TestExportDiffExcel:
    """Excel Diff 导出测试"""

    def _get_export_func(self):
        from report_exporter import export_diff_excel
        return export_diff_excel

    def test_creates_valid_xlsx(self, sample_comparison_result, sample_summary,
                                 sample_word_text, sample_pdf_text):
        export_func = self._get_export_func()
        result = export_func(sample_comparison_result, sample_summary,
                              sample_word_text, sample_pdf_text)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_contains_expected_sheets(self, sample_comparison_result, sample_summary,
                                       sample_word_text, sample_pdf_text):
        export_func = self._get_export_func()
        result = export_func(sample_comparison_result, sample_summary,
                              sample_word_text, sample_pdf_text)
        # Load with openpyxl to check sheets
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(result))
        sheet_names = wb.sheetnames
        assert "差异摘要" in sheet_names
        # Check that category sheets exist
        assert any("金额" in s for s in sheet_names)

    def test_empty_data_no_crash(self, empty_comparison_result, empty_summary):
        export_func = self._get_export_func()
        result = export_func(empty_comparison_result, empty_summary, "", "")
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_returns_bytes_object(self, sample_comparison_result, sample_summary,
                                   sample_word_text, sample_pdf_text):
        export_func = self._get_export_func()
        result = export_func(sample_comparison_result, sample_summary,
                              sample_word_text, sample_pdf_text)
        assert isinstance(result, bytes)


# ====================== PDF 导出测试 ======================

@pytest.mark.skipif(not REPORTLAB_AVAILABLE, reason="reportlab not installed")
class TestExportPdfReport:
    """PDF Report 导出测试"""

    def _get_export_func(self):
        from report_exporter import export_pdf_report
        return export_pdf_report

    def test_creates_valid_pdf(self, sample_comparison_result, sample_summary,
                                sample_word_text, sample_pdf_text):
        export_func = self._get_export_func()
        result = export_func(sample_comparison_result, sample_summary,
                              sample_word_text, sample_pdf_text)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_pdf_starts_with_magic(self, sample_comparison_result, sample_summary,
                                    sample_word_text, sample_pdf_text):
        export_func = self._get_export_func()
        result = export_func(sample_comparison_result, sample_summary,
                              sample_word_text, sample_pdf_text)
        # PDF should start with %PDF
        assert result[:4] == b"%PDF"

    def test_empty_data_no_crash(self, empty_comparison_result, empty_summary):
        export_func = self._get_export_func()
        result = export_func(empty_comparison_result, empty_summary, "", "")
        assert isinstance(result, bytes)
        assert result[:4] == b"%PDF"

    def test_returns_bytes_object(self, sample_comparison_result, sample_summary,
                                   sample_word_text, sample_pdf_text):
        export_func = self._get_export_func()
        result = export_func(sample_comparison_result, sample_summary,
                              sample_word_text, sample_pdf_text)
        assert isinstance(result, bytes)


# ====================== ZIP 全量导出测试 ======================

class TestExportFullPackage:
    """ZIP 全量导出测试"""

    def _get_export_func(self):
        from report_exporter import export_full_package
        return export_full_package

    def test_creates_valid_zip(self, sample_comparison_result, sample_summary,
                                sample_word_text, sample_pdf_text):
        export_func = self._get_export_func()
        result = export_func(sample_comparison_result, sample_summary,
                              sample_word_text, sample_pdf_text)
        assert isinstance(result, bytes)
        assert len(result) > 0
        assert zipfile.is_zipfile(io.BytesIO(result))

    def test_zip_contains_text_files(self, sample_comparison_result, sample_summary,
                                      sample_word_text, sample_pdf_text):
        export_func = self._get_export_func()
        result = export_func(sample_comparison_result, sample_summary,
                              sample_word_text, sample_pdf_text)
        zf = zipfile.ZipFile(io.BytesIO(result))
        names = zf.namelist()
        assert any("word_原文.txt" in n for n in names)
        assert any("pdf_ocr文本.txt" in n for n in names)

    def test_zip_text_content_correct(self, sample_comparison_result, sample_summary,
                                       sample_word_text, sample_pdf_text):
        export_func = self._get_export_func()
        result = export_func(sample_comparison_result, sample_summary,
                              sample_word_text, sample_pdf_text)
        zf = zipfile.ZipFile(io.BytesIO(result))
        for name in zf.namelist():
            if "word_原文.txt" in name:
                content = zf.read(name).decode("utf-8")
                assert content == sample_word_text
            elif "pdf_ocr文本.txt" in name:
                content = zf.read(name).decode("utf-8")
                assert content == sample_pdf_text

    def test_empty_data_no_crash(self, empty_comparison_result, empty_summary):
        export_func = self._get_export_func()
        result = export_func(empty_comparison_result, empty_summary, "", "")
        assert isinstance(result, bytes)