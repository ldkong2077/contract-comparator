"""
ExcelComparator 单元测试
测试 ExcelParser 解析、ExcelComparator 比对、差异报告导出
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import pytest

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from excel_comparator import ExcelParser, ExcelComparator, generate_diff_excel


# ============================================================
# 辅助函数：创建临时 Excel 文件
# ============================================================

def _create_simple_excel(path, data, sheet_name="Sheet1"):
    """
    创建简单 Excel 文件

    Args:
        path: 输出路径
        data: 二维列表，第一行为表头
        sheet_name: 工作表名称
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in data:
        ws.append(row)
    wb.save(path)
    wb.close()


def _create_merged_excel(path):
    """创建含合并单元格的 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "合并测试"
    # 表头
    ws.append(["名称", "类别", "金额"])
    # 合并单元格：A2:A3 合并
    ws.merge_cells("A2:A3")
    ws["A2"] = "项目A"
    ws["B2"] = "类别1"
    ws["C2"] = 1000
    ws["B3"] = "类别2"
    ws["C3"] = 2000
    wb.save(path)
    wb.close()


# ============================================================
# ExcelParser 测试
# ============================================================

class TestExcelParser:
    """Excel 解析器测试"""

    def test_excel_parser_basic(self, tmp_path):
        """基本解析：应正确提取表头和数据行"""
        file_path = str(tmp_path / "basic.xlsx")
        _create_simple_excel(file_path, [
            ["名称", "数量", "单价"],
            ["苹果", 10, 5.5],
            ["香蕉", 20, 3.2],
        ])

        parser = ExcelParser(file_path)
        result = parser.parse()

        assert "sheets" in result
        assert "full_text" in result
        assert len(result["sheets"]) >= 1

        sheet = result["sheets"][0]
        assert sheet["name"] == "Sheet1"
        assert sheet["row_count"] == 3
        assert sheet["headers"] == ["名称", "数量", "单价"]

    def test_excel_parser_file_not_found(self):
        """文件不存在应抛出 FileNotFoundError"""
        parser = ExcelParser("/nonexistent/file.xlsx")
        with pytest.raises(FileNotFoundError):
            parser.parse()

    def test_excel_parser_unsupported_format(self, tmp_path):
        """非 xlsx 格式应抛出 ValueError"""
        file_path = str(tmp_path / "test.csv")
        with open(file_path, "w") as f:
            f.write("a,b,c")
        parser = ExcelParser(file_path)
        with pytest.raises(ValueError, match="不支持的文件格式"):
            parser.parse()

    def test_excel_parser_merged_cells(self, tmp_path):
        """合并单元格应被正确填充"""
        file_path = str(tmp_path / "merged.xlsx")
        _create_merged_excel(file_path)

        parser = ExcelParser(file_path)
        result = parser.parse()

        sheet = result["sheets"][0]
        # 合并单元格的值应填充到所有子单元格
        rows = sheet["rows"]
        # A2 和 A3 应该都有值 "项目A"
        assert rows[1][0] == "项目A"
        assert rows[2][0] == "项目A"

    def test_excel_parser_empty_sheet(self, tmp_path):
        """空工作表应正常解析"""
        file_path = str(tmp_path / "empty.xlsx")
        wb = Workbook()
        wb.save(file_path)
        wb.close()

        parser = ExcelParser(file_path)
        result = parser.parse()
        assert len(result["sheets"]) >= 1

    def test_excel_parser_multiple_sheets(self, tmp_path):
        """多工作表应全部解析"""
        file_path = str(tmp_path / "multi_sheet.xlsx")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "表1"
        ws1.append(["A", "B"])
        ws1.append([1, 2])
        ws2 = wb.create_sheet("表2")
        ws2.append(["C", "D"])
        ws2.append([3, 4])
        wb.save(file_path)
        wb.close()

        parser = ExcelParser(file_path)
        result = parser.parse()
        assert len(result["sheets"]) == 2
        sheet_names = [s["name"] for s in result["sheets"]]
        assert "表1" in sheet_names
        assert "表2" in sheet_names


# ============================================================
# ExcelComparator 测试
# ============================================================

class TestExcelComparator:
    """Excel 比对引擎测试"""

    def test_excel_comparator_identical_files(self, tmp_path):
        """完全相同的文件应无差异"""
        data = [
            ["名称", "数量", "单价"],
            ["苹果", 10, 5.5],
            ["香蕉", 20, 3.2],
        ]
        file_a = str(tmp_path / "a.xlsx")
        file_b = str(tmp_path / "b.xlsx")
        _create_simple_excel(file_a, data)
        _create_simple_excel(file_b, data)

        comparator = ExcelComparator()
        result = comparator.compare(file_a, file_b)

        assert result["summary"]["total_differences"] == 0
        assert result["summary"]["has_critical_diff"] is False

    def test_excel_comparator_different_values(self, tmp_path):
        """不同数值应产生差异"""
        _create_simple_excel(str(tmp_path / "a.xlsx"), [
            ["名称", "数量", "金额"],
            ["苹果", 10, 1000],
        ])
        _create_simple_excel(str(tmp_path / "b.xlsx"), [
            ["名称", "数量", "金额"],
            ["苹果", 10, 2000],
        ])

        comparator = ExcelComparator()
        result = comparator.compare(
            str(tmp_path / "a.xlsx"),
            str(tmp_path / "b.xlsx"),
        )

        assert result["summary"]["total_differences"] > 0

    def test_excel_comparator_added_rows(self, tmp_path):
        """新增行应被检测到"""
        _create_simple_excel(str(tmp_path / "a.xlsx"), [
            ["名称", "数量"],
            ["苹果", 10],
        ])
        _create_simple_excel(str(tmp_path / "b.xlsx"), [
            ["名称", "数量"],
            ["苹果", 10],
            ["香蕉", 20],
        ])

        comparator = ExcelComparator()
        result = comparator.compare(
            str(tmp_path / "a.xlsx"),
            str(tmp_path / "b.xlsx"),
        )

        assert result["summary"]["total_differences"] > 0

    def test_excel_comparator_numeric_tolerance(self, tmp_path):
        """数值在容差范围内应视为相同"""
        _create_simple_excel(str(tmp_path / "a.xlsx"), [
            ["名称", "数值"],
            ["测试", 100.001],
        ])
        _create_simple_excel(str(tmp_path / "b.xlsx"), [
            ["名称", "数值"],
            ["测试", 100.002],
        ])

        # 容差 0.01，差值 0.001 在容差范围内
        comparator = ExcelComparator(tolerance=0.01)
        result = comparator.compare(
            str(tmp_path / "a.xlsx"),
            str(tmp_path / "b.xlsx"),
        )

        # 数值差异在容差内，应无差异
        assert result["summary"]["total_differences"] == 0

    def test_excel_comparator_only_in_a(self, tmp_path):
        """仅文件A有的工作表应被标记"""
        wb_a = Workbook()
        ws = wb_a.active
        ws.title = "仅A有"
        ws.append(["数据"])
        ws.append([1])
        file_a = str(tmp_path / "a.xlsx")
        wb_a.save(file_a)
        wb_a.close()

        wb_b = Workbook()
        ws_b = wb_b.active
        ws_b.title = "仅B有"
        ws_b.append(["数据"])
        ws_b.append([2])
        file_b = str(tmp_path / "b.xlsx")
        wb_b.save(file_b)
        wb_b.close()

        comparator = ExcelComparator()
        result = comparator.compare(file_a, file_b)

        # 两个工作表名称不同，应各标记为 only_in_a / only_in_b
        sheet_statuses = [s["status"] for s in result["sheets"]]
        assert "only_in_a" in sheet_statuses or "only_in_b" in sheet_statuses

    def test_excel_comparator_financial_column_high_risk(self, tmp_path):
        """金额列数值差异应为高风险"""
        _create_simple_excel(str(tmp_path / "a.xlsx"), [
            ["项目", "金额"],
            ["合同A", 100000],
        ])
        _create_simple_excel(str(tmp_path / "b.xlsx"), [
            ["项目", "金额"],
            ["合同A", 200000],
        ])

        comparator = ExcelComparator()
        result = comparator.compare(
            str(tmp_path / "a.xlsx"),
            str(tmp_path / "b.xlsx"),
        )

        # 金额列差异应为高风险
        has_high_risk = False
        for sheet_result in result["sheets"]:
            for diff in sheet_result.get("differences", []):
                if diff.get("risk") == "high":
                    has_high_risk = True
        assert has_high_risk or result["summary"]["has_critical_diff"]


# ============================================================
# 差异报告导出测试
# ============================================================

class TestGenerateDiffExcel:
    """差异报告 Excel 导出测试"""

    def test_generate_diff_excel(self, tmp_path):
        """应生成差异报告 Excel 文件"""
        result = {
            "summary": {
                "total_sheets_compared": 1,
                "total_differences": 2,
                "has_critical_diff": True,
                "sheet_summaries": [
                    {
                        "sheet_name": "Sheet1",
                        "status": "matched",
                        "total_diffs": 2,
                        "has_high_risk": True,
                    }
                ],
            },
            "sheets": [
                {
                    "sheet_name": "Sheet1",
                    "status": "matched",
                    "differences": [
                        {
                            "type": "cell_changed",
                            "row": 2,
                            "col": 3,
                            "col_name": "金额",
                            "value_a": 100000,
                            "value_b": 200000,
                            "risk": "high",
                            "category": "number",
                        },
                        {
                            "type": "row_added",
                            "row": 4,
                            "col": -1,
                            "col_name": "",
                            "value_a": None,
                            "value_b": ["新增行", 500],
                            "risk": "medium",
                            "category": "text",
                        },
                    ],
                    "stats": {
                        "added_rows": 1,
                        "deleted_rows": 0,
                        "modified_cells": 1,
                    },
                }
            ],
        }

        output_path = str(tmp_path / "diff_report.xlsx")
        abs_path = generate_diff_excel(result, output_path)

        assert os.path.exists(abs_path)
        # 验证文件可被 openpyxl 读取
        from openpyxl import load_workbook
        wb = load_workbook(abs_path)
        assert "比对摘要" in wb.sheetnames
        wb.close()

    def test_generate_diff_excel_only_in_a(self, tmp_path):
        """仅文件A有的工作表应正确导出"""
        result = {
            "summary": {
                "total_sheets_compared": 1,
                "total_differences": 1,
                "has_critical_diff": True,
                "sheet_summaries": [
                    {"sheet_name": "仅A", "status": "only_in_a", "total_diffs": 1, "has_high_risk": True}
                ],
            },
            "sheets": [
                {
                    "sheet_name": "仅A",
                    "status": "only_in_a",
                    "differences": [],
                    "stats": {"added_rows": 0, "deleted_rows": 1, "modified_cells": 0},
                }
            ],
        }

        output_path = str(tmp_path / "diff_only_a.xlsx")
        abs_path = generate_diff_excel(result, output_path)
        assert os.path.exists(abs_path)
