"""
Excel 电子表格比对模块
支持 .xlsx 格式的 Excel 文件解析、智能比对和差异报告导出
"""

import os
import re
import logging
from datetime import datetime
from difflib import SequenceMatcher
from typing import Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ============================================================
# ExcelParser - Excel 文件解析器
# ============================================================

class ExcelParser:
    """
    Excel 文件解析器
    支持 .xlsx 格式，处理合并单元格、公式缓存值、隐藏表等
    """

    # 大文件阈值（行数），超过此值使用流式读取
    LARGE_FILE_THRESHOLD = 5000

    def __init__(self, file_path: str, include_hidden: bool = False):
        """
        初始化解析器

        Args:
            file_path: Excel 文件路径（支持中文路径）
            include_hidden: 是否包含隐藏工作表，默认跳过
        """
        self.file_path = file_path
        self.include_hidden = include_hidden

    def parse(self) -> dict:
        """
        解析 Excel 文件，提取所有工作表的结构化数据

        Returns:
            {
                "sheets": [
                    {
                        "name": str,
                        "rows": list[list],
                        "headers": list,
                        "row_count": int,
                        "col_count": int
                    }
                ],
                "full_text": str
            }
        """
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"文件不存在: {self.file_path}")

        ext = os.path.splitext(self.file_path)[1].lower()
        if ext != ".xlsx":
            raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx")

        try:
            # read_only=True 适合大文件流式读取
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
        except Exception as e:
            logger.error(f"打开 Excel 文件失败: {self.file_path}, 错误: {e}")
            raise

        sheets = []
        full_text_parts = []

        try:
            for ws in wb.worksheets:
                # 跳过隐藏工作表
                if not self.include_hidden and ws.sheet_state != "visible":
                    logger.info(f"跳过隐藏工作表: {ws.title}")
                    continue

                sheet_data = self._parse_sheet(ws)
                sheets.append(sheet_data)

                # 拼接全文（用于全文检索）
                if sheet_data["full_text"]:
                    full_text_parts.append(
                        f"=== 工作表: {sheet_data['name']} ===\n"
                        + sheet_data["full_text"]
                    )
        finally:
            wb.close()

        # 如果 read_only 模式无法获取合并单元格信息，重新用普通模式读取
        # 因为 read_only 模式下 merged_cells 不可用
        sheets = self._fill_merged_cells_if_needed(sheets)

        return {
            "sheets": sheets,
            "full_text": "\n\n".join(full_text_parts),
        }

    def _parse_sheet(self, ws) -> dict:
        """
        解析单个工作表

        Args:
            ws: openpyxl 工作表对象

        Returns:
            工作表数据字典
        """
        rows = []
        row_count = 0
        col_count = 0

        # 流式逐行读取
        for row in ws.iter_rows(values_only=True):
            # 将 None 转为空字符串，保持列对齐
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                else:
                    row_data.append(cell)
            rows.append(row_data)
            row_count += 1
            # 取最大列数
            col_count = max(col_count, len(row_data))

        # 统一每行的列数（补齐空列）
        for i, row in enumerate(rows):
            if len(row) < col_count:
                rows[i] = row + [""] * (col_count - len(row))

        # 提取表头（第一行非空行）
        headers = self._extract_headers(rows)

        # 生成全文文本
        full_text = self._rows_to_text(rows, headers)

        return {
            "name": ws.title,
            "rows": rows,
            "headers": headers,
            "row_count": row_count,
            "col_count": col_count,
            "full_text": full_text,
        }

    def _extract_headers(self, rows: list) -> list:
        """
        提取表头行（取第一个非空行作为表头）

        Args:
            rows: 所有行数据

        Returns:
            表头列表
        """
        for row in rows:
            # 如果行中有非空值，视为表头
            if any(cell for cell in row if cell != ""):
                return [str(cell) if cell != "" else f"列{i+1}" for i, cell in enumerate(row)]
        # 全部为空，返回默认列名
        if rows:
            return [f"列{i+1}" for i in range(len(rows[0]))]
        return []

    def _rows_to_text(self, rows: list, headers: list) -> str:
        """
        将行数据转换为文本（用于全文检索）

        Args:
            rows: 行数据
            headers: 表头

        Returns:
            文本字符串
        """
        lines = []
        for row_idx, row in enumerate(rows):
            # 跳过全空行
            if not any(cell for cell in row if cell != ""):
                continue
            parts = []
            for col_idx, cell in enumerate(row):
                if cell != "":
                    col_name = headers[col_idx] if col_idx < len(headers) else f"列{col_idx+1}"
                    parts.append(f"{col_name}: {cell}")
            if parts:
                lines.append(f"行{row_idx+1}: " + " | ".join(parts))
        return "\n".join(lines)

    def _fill_merged_cells_if_needed(self, sheets: list) -> list:
        """
        处理合并单元格：重新用普通模式打开文件，将合并单元格的值填充到所有子单元格

        Args:
            sheets: 已解析的工作表列表

        Returns:
            更新后的工作表列表
        """
        try:
            wb = load_workbook(self.file_path, data_only=True)
        except Exception as e:
            logger.warning(f"重新打开文件处理合并单元格失败: {e}")
            return sheets

        try:
            for sheet_data in sheets:
                ws = None
                for s in wb.worksheets:
                    if s.title == sheet_data["name"]:
                        ws = s
                        break
                if ws is None:
                    continue

                # 获取合并单元格范围
                merged_ranges = list(ws.merged_cells.ranges)
                if not merged_ranges:
                    continue

                # 构建合并单元格映射：{(row, col): value}
                merge_map = {}
                for merged_range in merged_ranges:
                    min_row = merged_range.min_row
                    min_col = merged_range.min_col
                    max_row = merged_range.max_row
                    max_col = merged_range.max_col

                    # 取左上角单元格的值
                    top_left_value = ws.cell(row=min_row, column=min_col).value
                    if top_left_value is None:
                        top_left_value = ""

                    # 将值填充到所有子单元格
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            merge_map[(r - 1, c - 1)] = top_left_value

                # 更新 rows 数据
                rows = sheet_data["rows"]
                for (r, c), value in merge_map.items():
                    if r < len(rows) and c < len(rows[r]):
                        rows[r][c] = value

                sheet_data["rows"] = rows

                # 重新生成全文
                sheet_data["full_text"] = self._rows_to_text(
                    rows, sheet_data["headers"]
                )
        finally:
            wb.close()

        return sheets


# ============================================================
# ExcelComparator - Excel 比对引擎
# ============================================================

class ExcelComparator:
    """
    Excel 电子表格比对引擎
    支持类型感知的单元格比对、智能列匹配、关键列检测、行匹配
    """

    # 金额/财务相关列名关键词
    FINANCIAL_KEYWORDS = [
        "金额", "价格", "费用", "总价", "合计", "总额", "总计",
        "单价", "价款", "报酬", "违约金", "赔偿金", "保证金",
        "押金", "罚金", "含税", "不含税", "增值税", "税额",
        "成本", "利润", "收入", "支出", "预算", "结算",
        "应收", "应付", "实付", "实收", "余额",
    ]

    # 日期相关列名关键词
    DATE_KEYWORDS = [
        "日期", "时间", "年月", "签订日", "生效日", "到期日",
        "起始日", "终止日", "创建时间", "修改时间", "更新时间",
        "开始日期", "结束日期", "截止日期", "交付日期",
    ]

    # ID/编号类列名关键词（用于关键列检测）
    ID_KEYWORDS = [
        "编号", "序号", "ID", "id", "No", "no", "NO",
        "合同号", "合同编号", "订单号", "订单编号",
        "项目编号", "项目号", "代码", "编码",
        "身份证", "工号", "学号", "卡号", "账号",
    ]

    def __init__(self, tolerance: float = 0.01):
        """
        初始化比对引擎

        Args:
            tolerance: 数值比较容差，默认 0.01
        """
        self.tolerance = tolerance
        self.parser = None  # 延迟初始化

    def compare(self, excel_a_path: str, excel_b_path: str) -> dict:
        """
        全量比对两个 Excel 文件

        Args:
            excel_a_path: 文件A路径（基准文件）
            excel_b_path: 文件B路径（对比文件）

        Returns:
            结构化比对结果
        """
        # 解析两个文件
        parser_a = ExcelParser(excel_a_path)
        parser_b = ExcelParser(excel_b_path)

        data_a = parser_a.parse()
        data_b = parser_b.parse()

        # 按工作表名称匹配
        sheets_a = {s["name"]: s for s in data_a["sheets"]}
        sheets_b = {s["name"]: s for s in data_b["sheets"]}

        all_sheet_names = list(dict.fromkeys(
            list(sheets_a.keys()) + list(sheets_b.keys())
        ))

        sheet_results = []
        total_differences = 0
        has_critical_diff = False
        sheet_summaries = []

        for name in all_sheet_names:
            if name in sheets_a and name in sheets_b:
                # 两个文件都有该工作表，执行比对
                result = self.compare_sheets(
                    sheets_a[name]["rows"],
                    sheets_b[name]["rows"],
                    sheets_a[name]["headers"],
                    sheets_b[name]["headers"],
                )
                diff_count = len(result["differences"])
                total_differences += diff_count

                # 检查是否有高风险差异
                has_high = any(d["risk"] == "high" for d in result["differences"])
                if has_high:
                    has_critical_diff = True

                sheet_results.append({
                    "sheet_name": name,
                    "status": "matched",
                    "differences": result["differences"],
                    "stats": result["stats"],
                })

                sheet_summaries.append({
                    "sheet_name": name,
                    "status": "matched",
                    "total_diffs": diff_count,
                    "has_high_risk": has_high,
                })

            elif name in sheets_a:
                # 仅文件A有该工作表
                row_count = sheets_a[name]["row_count"]
                sheet_results.append({
                    "sheet_name": name,
                    "status": "only_in_a",
                    "differences": [],
                    "stats": {
                        "added_rows": 0,
                        "deleted_rows": row_count,
                        "modified_cells": 0,
                    },
                })
                total_differences += row_count

                sheet_summaries.append({
                    "sheet_name": name,
                    "status": "only_in_a",
                    "total_diffs": row_count,
                    "has_high_risk": True,
                })
                has_critical_diff = True

            else:
                # 仅文件B有该工作表
                row_count = sheets_b[name]["row_count"]
                sheet_results.append({
                    "sheet_name": name,
                    "status": "only_in_b",
                    "differences": [],
                    "stats": {
                        "added_rows": row_count,
                        "deleted_rows": 0,
                        "modified_cells": 0,
                    },
                })
                total_differences += row_count

                sheet_summaries.append({
                    "sheet_name": name,
                    "status": "only_in_b",
                    "total_diffs": row_count,
                    "has_high_risk": True,
                })
                has_critical_diff = True

        return {
            "summary": {
                "total_sheets_compared": len(all_sheet_names),
                "total_differences": total_differences,
                "has_critical_diff": has_critical_diff,
                "sheet_summaries": sheet_summaries,
            },
            "sheets": sheet_results,
        }

    def compare_sheets(
        self,
        sheet_a: list,
        sheet_b: list,
        headers_a: list,
        headers_b: list,
    ) -> dict:
        """
        逐单元格比对两个工作表

        Args:
            sheet_a: 工作表A的行数据
            sheet_b: 工作表B的行数据
            headers_a: 工作表A的表头
            headers_b: 工作表B的表头

        Returns:
            {
                "differences": [...],
                "stats": {"added_rows": int, "deleted_rows": int, "modified_cells": int}
            }
        """
        differences = []

        # 智能列匹配
        col_mapping = self._match_columns(headers_a, headers_b)

        # 检测关键列
        key_cols_a = self._detect_key_columns(headers_a, sheet_a)

        # 行匹配：优先使用关键列匹配，否则位置匹配
        row_mapping = self._match_rows(
            sheet_a, sheet_b, headers_a, headers_b, col_mapping, key_cols_a
        )

        # 统计
        added_rows = 0
        deleted_rows = 0
        modified_cells = 0

        matched_a_indices = set()
        matched_b_indices = set()

        # 比对已匹配的行
        for idx_a, idx_b in row_mapping:
            matched_a_indices.add(idx_a)
            matched_b_indices.add(idx_b)

            # 跳过表头行（第一行）
            row_a = sheet_a[idx_a]
            row_b = sheet_b[idx_b]

            # 按列映射逐单元格比较
            for col_a, col_b in col_mapping:
                val_a = row_a[col_a] if col_a < len(row_a) else ""
                val_b = row_b[col_b] if col_b < len(row_b) else ""

                col_name = headers_a[col_a] if col_a < len(headers_a) else f"列{col_a+1}"

                diff = self._compare_cell(val_a, val_b, idx_a, col_a, col_name)
                if diff:
                    differences.append(diff)
                    modified_cells += 1

        # 检测删除的行（在A中但未匹配到B）
        for i in range(len(sheet_a)):
            if i not in matched_a_indices:
                # 跳过全空行
                if not any(cell for cell in sheet_a[i] if cell != ""):
                    continue
                differences.append({
                    "type": "row_deleted",
                    "row": i + 1,
                    "col": -1,
                    "col_name": "",
                    "value_a": sheet_a[i],
                    "value_b": None,
                    "risk": "medium",
                    "category": "text",
                })
                deleted_rows += 1

        # 检测新增的行（在B中但未匹配到A）
        for j in range(len(sheet_b)):
            if j not in matched_b_indices:
                # 跳过全空行
                if not any(cell for cell in sheet_b[j] if cell != ""):
                    continue
                differences.append({
                    "type": "row_added",
                    "row": j + 1,
                    "col": -1,
                    "col_name": "",
                    "value_a": None,
                    "value_b": sheet_b[j],
                    "risk": "medium",
                    "category": "text",
                })
                added_rows += 1

        # 检测行移动（删除行和新增行中内容相似的配对）
        self._detect_moved_rows(differences, added_rows, deleted_rows)

        return {
            "differences": differences,
            "stats": {
                "added_rows": added_rows,
                "deleted_rows": deleted_rows,
                "modified_cells": modified_cells,
            },
        }

    def _compare_cell(
        self, val_a, val_b, row: int, col: int, col_name: str
    ) -> Optional[dict]:
        """
        比较两个单元格的值

        Args:
            val_a: 单元格A的值
            val_b: 单元格B的值
            row: 行号
            col: 列号
            col_name: 列名

        Returns:
            差异字典，如果相同则返回 None
        """
        # 归一化处理
        norm_a = self._normalize_value(val_a)
        norm_b = self._normalize_value(val_b)

        # 完全相同则无差异
        if norm_a == norm_b:
            return None

        # 判断值类型和风险等级
        category = self._classify_category(val_a, val_b, col_name)
        risk = self._classify_risk(category, col_name)

        # 数值类型使用容差比较
        if category == "number":
            num_a = self._to_number(val_a)
            num_b = self._to_number(val_b)
            if num_a is not None and num_b is not None:
                if abs(num_a - num_b) <= self.tolerance:
                    return None  # 在容差范围内，视为相同

        return {
            "type": "cell_changed",
            "row": row + 1,
            "col": col + 1,
            "col_name": col_name,
            "value_a": val_a,
            "value_b": val_b,
            "risk": risk,
            "category": category,
        }

    def _normalize_value(self, value) -> str:
        """
        归一化值：去除首尾空白、统一空白字符、统一大小写

        Args:
            value: 原始值

        Returns:
            归一化后的字符串
        """
        if value is None:
            return ""
        s = str(value).strip()
        # 将连续空白替换为单个空格
        s = re.sub(r'\s+', ' ', s)
        # 全角空格转半角
        s = s.replace('\u3000', ' ')
        return s

    def _to_number(self, value) -> Optional[float]:
        """
        尝试将值转换为数字

        Args:
            value: 原始值

        Returns:
            浮点数或 None
        """
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, datetime):
            return None
        s = str(value).strip()
        # 去除千分位逗号
        s = s.replace(",", "")
        # 去除货币符号
        s = re.sub(r'[¥￥$€]', '', s)
        # 去除百分号并转换
        if s.endswith('%'):
            try:
                return float(s[:-1]) / 100.0
            except ValueError:
                return None
        try:
            return float(s)
        except (ValueError, TypeError):
            return None

    def _classify_category(self, val_a, val_b, col_name: str) -> str:
        """
        分类差异类型：number / date / text / formula

        Args:
            val_a: 值A
            val_b: 值B
            col_name: 列名

        Returns:
            类别字符串
        """
        # 检查是否为日期类型
        if isinstance(val_a, datetime) or isinstance(val_b, datetime):
            return "date"

        # 检查列名是否暗示日期
        col_lower = col_name.lower()
        for kw in self.DATE_KEYWORDS:
            if kw in col_lower:
                # 验证值是否看起来像日期
                if self._looks_like_date(val_a) or self._looks_like_date(val_b):
                    return "date"

        # 检查是否为数字类型
        num_a = self._to_number(val_a)
        num_b = self._to_number(val_b)
        if num_a is not None and num_b is not None:
            return "number"
        # 其中一个是数字，另一个是空或非数字但列名暗示数值
        if (num_a is not None or num_b is not None):
            for kw in self.FINANCIAL_KEYWORDS:
                if kw in col_lower:
                    return "number"

        # 检查是否为公式（以 = 开头的字符串）
        str_a = str(val_a) if val_a else ""
        str_b = str(val_b) if val_b else ""
        if str_a.startswith("=") or str_b.startswith("="):
            return "formula"

        return "text"

    def _looks_like_date(self, value) -> bool:
        """
        判断值是否看起来像日期

        Args:
            value: 待判断的值

        Returns:
            是否像日期
        """
        if isinstance(value, datetime):
            return True
        s = str(value).strip()
        # 常见日期格式
        date_patterns = [
            r'^\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日号]?$',
            r'^\d{4}年\s*\d{1,2}月\s*\d{1,2}日$',
            r'^\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}$',
        ]
        for pattern in date_patterns:
            if re.match(pattern, s):
                return True
        return False

    def _classify_risk(self, category: str, col_name: str) -> str:
        """
        风险等级分类

        - high: 金额/财务列的数值差异、日期变更
        - medium: 文本内容变更、公式变更
        - low: 格式差异、空白差异

        Args:
            category: 差异类别
            col_name: 列名

        Returns:
            风险等级 "high" / "medium" / "low"
        """
        col_lower = col_name.lower()

        # 金额/财务列的数值差异 → 高风险
        if category == "number":
            for kw in self.FINANCIAL_KEYWORDS:
                if kw in col_lower:
                    return "high"
            # 普通数字列 → 中风险
            return "medium"

        # 日期变更 → 高风险
        if category == "date":
            return "high"

        # 公式变更 → 中风险
        if category == "formula":
            return "medium"

        # 文本变更 → 中风险
        if category == "text":
            # 检查是否仅为空白差异
            return "medium"

        return "low"

    # ============================================================
    # 智能列匹配
    # ============================================================

    def _match_columns(self, headers_a: list, headers_b: list) -> list:
        """
        智能列匹配：通过表头名称匹配列，支持模糊匹配和列重排序

        Args:
            headers_a: 工作表A的表头
            headers_b: 工作表B的表头

        Returns:
            列映射列表 [(col_a_idx, col_b_idx), ...]
        """
        mapping = []
        matched_b = set()

        # 第一轮：精确匹配（归一化后）
        norm_a = [self._normalize_header(h) for h in headers_a]
        norm_b = [self._normalize_header(h) for h in headers_b]

        for i, ha in enumerate(norm_a):
            for j, hb in enumerate(norm_b):
                if j in matched_b:
                    continue
                if ha and hb and ha == hb:
                    mapping.append((i, j))
                    matched_b.add(j)
                    break

        # 第二轮：模糊匹配（相似度 > 0.7）
        unmatched_a = [i for i in range(len(headers_a)) if not any(m[0] == i for m in mapping)]
        unmatched_b = [j for j in range(len(headers_b)) if j not in matched_b]

        for i in unmatched_a:
            best_j = -1
            best_score = 0.7  # 最低相似度阈值

            for j in unmatched_b:
                score = self._header_similarity(headers_a[i], headers_b[j])
                if score > best_score:
                    best_score = score
                    best_j = j

            if best_j >= 0:
                mapping.append((i, best_j))
                matched_b.add(best_j)
                unmatched_b.remove(best_j)

        # 第三轮：位置匹配（未匹配的列按位置对应）
        mapped_a = {m[0] for m in mapping}
        mapped_b = {m[1] for m in mapping}

        pos_unmatched_a = sorted(i for i in range(len(headers_a)) if i not in mapped_a)
        pos_unmatched_b = sorted(j for j in range(len(headers_b)) if j not in mapped_b)

        for i, j in zip(pos_unmatched_a, pos_unmatched_b):
            mapping.append((i, j))

        return mapping

    def _normalize_header(self, header: str) -> str:
        """
        归一化表头名称：去除空白、标点，统一大小写

        Args:
            header: 原始表头

        Returns:
            归一化后的表头
        """
        if not header:
            return ""
        s = str(header).strip()
        # 去除标点和空白
        s = re.sub(r'[\s\-_·•·:：()（）\[\]【】]', '', s)
        return s.lower()

    def _header_similarity(self, h1: str, h2: str) -> float:
        """
        计算两个表头的相似度

        Args:
            h1: 表头1
            h2: 表头2

        Returns:
            相似度 0~1
        """
        n1 = self._normalize_header(h1)
        n2 = self._normalize_header(h2)
        if not n1 or not n2:
            return 0.0
        return SequenceMatcher(None, n1, n2).ratio()

    # ============================================================
    # 关键列检测
    # ============================================================

    def _detect_key_columns(self, headers: list, rows: list) -> list:
        """
        自动检测关键列（ID/编号类列）
        关键列特征：列名包含ID关键词 + 值具有唯一性

        Args:
            headers: 表头列表
            rows: 行数据

        Returns:
            关键列索引列表
        """
        key_cols = []

        for col_idx, header in enumerate(headers):
            header_str = str(header).lower()

            # 检查列名是否包含ID关键词
            is_id_col = any(kw.lower() in header_str for kw in self.ID_KEYWORDS)

            if is_id_col:
                # 验证值的唯一性（非空值的唯一率 > 80%）
                values = []
                for row in rows[1:]:  # 跳过表头行
                    if col_idx < len(row):
                        val = row[col_idx]
                        if val != "":
                            values.append(str(val))

                if values:
                    unique_ratio = len(set(values)) / len(values)
                    if unique_ratio >= 0.8:
                        key_cols.append(col_idx)

        return key_cols

    # ============================================================
    # 行匹配
    # ============================================================

    def _match_rows(
        self,
        sheet_a: list,
        sheet_b: list,
        headers_a: list,
        headers_b: list,
        col_mapping: list,
        key_cols_a: list,
    ) -> list:
        """
        行匹配：优先使用关键列匹配，否则使用位置匹配

        Args:
            sheet_a: 工作表A行数据
            sheet_b: 工作表B行数据
            headers_a: 工作表A表头
            headers_b: 工作表B表头
            col_mapping: 列映射
            key_cols_a: 关键列索引

        Returns:
            行映射列表 [(row_a_idx, row_b_idx), ...]
        """
        # 如果有关键列，使用关键列匹配
        if key_cols_a:
            row_mapping = self._match_rows_by_key(
                sheet_a, sheet_b, col_mapping, key_cols_a
            )
            if row_mapping:
                return row_mapping

        # 退回位置匹配
        return self._match_rows_by_position(sheet_a, sheet_b)

    def _match_rows_by_key(
        self,
        sheet_a: list,
        sheet_b: list,
        col_mapping: list,
        key_cols_a: list,
    ) -> list:
        """
        使用关键列值进行行匹配

        Args:
            sheet_a: 工作表A行数据
            sheet_b: 工作表B行数据
            col_mapping: 列映射
            key_cols_a: 关键列索引（A中的）

        Returns:
            行映射列表
        """
        # 找到关键列在B中的对应列
        key_col_mapping = {}
        for col_a in key_cols_a:
            for ca, cb in col_mapping:
                if ca == col_a:
                    key_col_mapping[col_a] = cb
                    break

        if not key_col_mapping:
            return []

        # 构建 B 中关键列值的索引
        b_key_index = {}  # {key_value: [row_indices]}
        for j, row_b in enumerate(sheet_b):
            # 跳过空行
            if not any(cell for cell in row_b if cell != ""):
                continue

            key_values = []
            for col_a, col_b in key_col_mapping.items():
                val = row_b[col_b] if col_b < len(row_b) else ""
                key_values.append(str(val).strip())
            key = "||".join(key_values)

            if key and key != "||".join([""] * len(key_col_mapping)):
                if key not in b_key_index:
                    b_key_index[key] = []
                b_key_index[key].append(j)

        # 匹配 A 中的行
        row_mapping = []
        matched_b = set()

        for i, row_a in enumerate(sheet_a):
            # 跳过空行
            if not any(cell for cell in row_a if cell != ""):
                continue

            key_values = []
            for col_a in key_col_mapping.keys():
                val = row_a[col_a] if col_a < len(row_a) else ""
                key_values.append(str(val).strip())
            key = "||".join(key_values)

            if key in b_key_index:
                for j in b_key_index[key]:
                    if j not in matched_b:
                        row_mapping.append((i, j))
                        matched_b.add(j)
                        break

        return row_mapping

    def _match_rows_by_position(self, sheet_a: list, sheet_b: list) -> list:
        """
        位置匹配：按行号一一对应

        Args:
            sheet_a: 工作表A行数据
            sheet_b: 工作表B行数据

        Returns:
            行映射列表
        """
        mapping = []
        max_rows = max(len(sheet_a), len(sheet_b))

        for i in range(max_rows):
            has_a = i < len(sheet_a) and any(cell for cell in sheet_a[i] if cell != "")
            has_b = i < len(sheet_b) and any(cell for cell in sheet_b[i] if cell != "")

            if has_a and has_b:
                mapping.append((i, i))

        return mapping

    def _detect_moved_rows(self, differences: list, added_rows: int, deleted_rows: int) -> None:
        """
        检测行移动：在新增行和删除行中寻找内容相似的配对

        Args:
            differences: 差异列表（会被原地修改）
            added_rows: 新增行数
            deleted_rows: 删除行数
        """
        deleted_diffs = [d for d in differences if d["type"] == "row_deleted"]
        added_diffs = [d for d in differences if d["type"] == "row_added"]

        if not deleted_diffs or not added_diffs:
            return

        moved_pairs = []
        used_deleted = set()
        used_added = set()

        for i, del_diff in enumerate(deleted_diffs):
            if i in used_deleted:
                continue
            val_a = del_diff["value_a"]
            if isinstance(val_a, list):
                text_a = " ".join(str(c) for c in val_a if c != "")
            else:
                text_a = str(val_a)

            best_j = -1
            best_score = 0.6  # 行移动相似度阈值

            for j, add_diff in enumerate(added_diffs):
                if j in used_added:
                    continue
                val_b = add_diff["value_b"]
                if isinstance(val_b, list):
                    text_b = " ".join(str(c) for c in val_b if c != "")
                else:
                    text_b = str(val_b)

                score = SequenceMatcher(None, text_a, text_b).ratio()
                if score > best_score:
                    best_score = score
                    best_j = j

            if best_j >= 0:
                moved_pairs.append((i, best_j))
                used_deleted.add(i)
                used_added.add(best_j)

        # 将匹配的行从 deleted/added 改为 moved
        for i, j in moved_pairs:
            del_diff = deleted_diffs[i]
            add_diff = added_diffs[j]

            # 创建 moved 类型的差异
            differences.append({
                "type": "row_moved",
                "row": del_diff["row"],
                "col": -1,
                "col_name": "",
                "value_a": del_diff["value_a"],
                "value_b": add_diff["value_b"],
                "risk": "low",
                "category": "text",
            })

        # 移除已标记为 moved 的原始差异
        moved_del_indices = {id(deleted_diffs[i]) for i, _ in moved_pairs}
        moved_add_indices = {id(added_diffs[j]) for _, j in moved_pairs}

        differences[:] = [
            d for d in differences
            if d["type"] != "row_deleted" or id(d) not in moved_del_indices
        ]
        differences[:] = [
            d for d in differences
            if d["type"] != "row_added" or id(d) not in moved_add_indices
        ]


# ============================================================
# 差异报告导出
# ============================================================

def generate_diff_excel(result: dict, output_path: str) -> str:
    """
    生成差异报告 Excel 文件
    颜色编码：红色=删除，绿色=新增，黄色=修改

    Args:
        result: ExcelComparator.compare() 的返回结果
        output_path: 输出文件路径

    Returns:
        输出文件的绝对路径
    """
    wb = Workbook()

    # ----- 通用样式定义 -----
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font = Font(name="Microsoft YaHei", size=10)
    body_alignment = Alignment(vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )

    # 差异类型对应的填充色
    diff_fills = {
        "cell_changed": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),  # 黄色-修改
        "row_added": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),      # 绿色-新增
        "row_deleted": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),     # 红色-删除
        "row_moved": PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid"),       # 紫色-移动
    }

    # 风险等级对应的字体颜色
    risk_fonts = {
        "high": Font(name="Microsoft YaHei", size=10, color="D32F2F", bold=True),
        "medium": Font(name="Microsoft YaHei", size=10, color="F57C00"),
        "low": Font(name="Microsoft YaHei", size=10, color="1976D2"),
    }

    risk_labels = {"high": "高风险", "medium": "中风险", "low": "低风险"}
    type_labels = {
        "cell_changed": "单元格变更",
        "row_added": "新增行",
        "row_deleted": "删除行",
        "row_moved": "行移动",
    }
    category_labels = {
        "number": "数值",
        "date": "日期",
        "text": "文本",
        "formula": "公式",
    }

    def _write_header(ws, headers: list, col_widths: list):
        """写入表头"""
        for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    def _format_value(val) -> str:
        """格式化差异值用于显示"""
        if val is None:
            return ""
        if isinstance(val, list):
            return " | ".join(str(c) for c in val if c != "")
        return str(val)

    # ===================== 摘要工作表 =====================
    ws_summary = wb.active
    ws_summary.title = "比对摘要"

    summary = result.get("summary", {})
    summary_headers = ["指标", "数值", "说明"]
    summary_widths = [20, 16, 50]
    _write_header(ws_summary, summary_headers, summary_widths)

    summary_rows = [
        ("比对工作表数", summary.get("total_sheets_compared", 0), ""),
        ("差异总数", summary.get("total_differences", 0), ""),
        (
            "是否存在高风险差异",
            "是" if summary.get("has_critical_diff") else "否",
            "高风险差异包括金额/财务列数值变更、日期变更，建议重点核查",
        ),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "自动生成"),
    ]

    # 各工作表摘要
    for ss in summary.get("sheet_summaries", []):
        status_label = {"matched": "已匹配", "only_in_a": "仅文件A", "only_in_b": "仅文件B"}.get(
            ss["status"], ss["status"]
        )
        summary_rows.append((
            f"  工作表: {ss['sheet_name']}",
            f"差异 {ss.get('total_diffs', 0)}",
            f"状态: {status_label}" + ("，含高风险" if ss.get("has_high_risk") else ""),
        ))

    for row_idx, (label, val, note) in enumerate(summary_rows, 2):
        for col_idx, v in enumerate([label, str(val), note], 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=v)
            cell.font = body_font
            cell.alignment = body_alignment
            cell.border = thin_border

    # ===================== 各工作表差异明细 =====================
    for sheet_result in result.get("sheets", []):
        sheet_name = sheet_result["sheet_name"]
        status = sheet_result["status"]
        differences = sheet_result.get("differences", [])
        stats = sheet_result.get("stats", {})

        # 工作表名称长度限制（Excel 工作表名最长31字符）
        tab_name = f"差异_{sheet_name}"[:31]

        ws = wb.create_sheet(title=tab_name)

        # 工作表状态说明
        if status == "only_in_a":
            ws.cell(row=1, column=1, value=f"工作表 '{sheet_name}' 仅存在于文件A中")
            ws.cell(row=1, column=1).font = Font(name="Microsoft YaHei", size=12, color="D32F2F", bold=True)
            continue
        elif status == "only_in_b":
            ws.cell(row=1, column=1, value=f"工作表 '{sheet_name}' 仅存在于文件B中")
            ws.cell(row=1, column=1).font = Font(name="Microsoft YaHei", size=12, color="1976D2", bold=True)
            continue

        # 差异明细表头
        detail_headers = ["行号", "列号", "列名", "差异类型", "风险等级", "类别", "文件A的值", "文件B的值"]
        detail_widths = [8, 8, 18, 14, 10, 10, 30, 30]
        _write_header(ws, detail_headers, detail_widths)

        # 统计信息行
        stats_row = 2
        stats_text = (
            f"新增行: {stats.get('added_rows', 0)} | "
            f"删除行: {stats.get('deleted_rows', 0)} | "
            f"修改单元格: {stats.get('modified_cells', 0)}"
        )
        cell = ws.cell(row=stats_row, column=1, value=stats_text)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="1A1A2E")
        ws.merge_cells(start_row=stats_row, start_column=1, end_row=stats_row, end_column=8)

        # 写入差异明细
        for diff_idx, diff in enumerate(differences):
            row_num = stats_row + 1 + diff_idx
            diff_type = diff.get("type", "")
            risk = diff.get("risk", "low")
            category = diff.get("category", "text")

            values = [
                diff.get("row", ""),
                diff.get("col", "") if diff.get("col", -1) >= 0 else "",
                diff.get("col_name", ""),
                type_labels.get(diff_type, diff_type),
                risk_labels.get(risk, risk),
                category_labels.get(category, category),
                _format_value(diff.get("value_a")),
                _format_value(diff.get("value_b")),
            ]

            # 获取差异类型对应的填充色
            fill = diff_fills.get(diff_type)
            font = risk_fonts.get(risk, body_font)

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = font if col_idx == 5 else body_font
                cell.alignment = body_alignment
                cell.border = thin_border
                if fill:
                    cell.fill = fill

    # 保存文件
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    return os.path.abspath(output_path)
